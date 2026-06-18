"""
批量导入服务模块

提供学生信息批量导入功能，支持 Excel (.xlsx) 和 CSV (.csv) 文件格式。
"""

import csv
import io
import json
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session

from src.models.student import Student
from src.models.exam_total import StudentExamTotal
from src.repositories.student_repo import StudentRepository
from src.core.exceptions import ValidationException


class ImportService:
    """批量导入服务类"""

    def __init__(self, db: Session):
        self.db = db
        self.student_repo = StudentRepository(db)

    def parse_csv_file(self, file_content: bytes) -> dict:
        """
        解析 CSV 文件内容
        
        Args:
            file_content: CSV 文件内容（字节）
            
        Returns:
            包含 students 和 errors 的字典
        """
        try:
            # 尝试不同编码，优先使用 utf-8-sig 处理 BOM
            content = None
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
                try:
                    content = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                raise ValidationException("无法识别文件编码，请使用 UTF-8 或 GBK 编码")

            # 去除可能残留的 BOM 字符
            if content.startswith('\ufeff'):
                content = content[1:]

            # 解析 CSV
            reader = csv.DictReader(io.StringIO(content))
            students = []
            errors = []
            
            for row_num, row in enumerate(reader, start=2):  # 从第2行开始（第1行是表头）
                # 清理字段名和值的空白字符
                cleaned_row = {k.strip(): v for k, v in row.items() if k is not None}
                result = self._validate_row(cleaned_row, row_num)
                if result['data']:
                    students.append(result['data'])
                if result['errors']:
                    errors.append({
                        'row': row_num,
                        'student_id': str(cleaned_row.get('学号', '')).strip() or '未知',
                        'error': '; '.join(result['errors'])
                    })
            
            return {'students': students, 'errors': errors}

        except csv.Error as e:
            raise ValidationException(f"CSV 文件格式错误: {str(e)}")

    def parse_excel_file(self, file_content: bytes) -> dict:
        """
        解析 Excel 文件内容
        
        Args:
            file_content: Excel 文件内容（字节）
            
        Returns:
            包含 students 和 errors 的字典
        """
        try:
            import openpyxl
        except ImportError:
            raise ValidationException("请安装 openpyxl 库以支持 Excel 文件")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active

            if ws is None:
                raise ValidationException("Excel 文件没有活动工作表")

            # 读取表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")

            # 验证表头
            required_headers = ['学号', '姓名', '性别', '班级', '入学年份']
            for header in required_headers:
                if header not in headers:
                    raise ValidationException(f"Excel 文件缺少必要列: {header}")

            # 解析数据行
            students = []
            errors = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                    
                row_dict = dict(zip(headers, row))
                result = self._validate_row(row_dict, row_num)
                if result['data']:
                    students.append(result['data'])
                if result['errors']:
                    errors.append({
                        'row': row_num,
                        'student_id': str(row_dict.get('学号', '')).strip() or '未知',
                        'error': '; '.join(result['errors'])
                    })

            wb.close()
            return {'students': students, 'errors': errors}

        except Exception as e:
            if isinstance(e, ValidationException):
                raise
            raise ValidationException(f"Excel 文件解析错误: {str(e)}")

    def _validate_row(self, row: dict, row_num: int) -> dict:
        """
        验证单行数据
        
        Args:
            row: 行数据字典
            row_num: 行号
            
        Returns:
            包含 data 和 errors 的字典：
            - data: 验证通过的学生数据（有错误时为 None）
            - errors: 错误信息列表
        """
        errors = []

        # 验证学号
        student_id = str(row.get('学号', '')).strip()
        if not student_id:
            errors.append("学号不能为空")
        elif len(student_id) != 8 or not student_id.isdigit():
            errors.append(f"学号格式错误，应为8位数字，当前值: {student_id}")
        elif int(student_id[:4]) < 2000 or int(student_id[:4]) > 2100:
            errors.append(f"学号年份部分应在2000-2100之间，当前值: {student_id[:4]}")

        # 验证姓名
        name = str(row.get('姓名', '')).strip()
        if not name:
            errors.append("姓名不能为空")
        elif len(name) < 2 or len(name) > 20:
            errors.append(f"姓名长度应在2-20个字符之间，当前长度: {len(name)}")

        # 验证性别
        gender = str(row.get('性别', '')).strip()
        if not gender:
            errors.append("性别不能为空")
        elif gender not in ['男', '女']:
            errors.append(f"性别只能是'男'或'女'，当前值: {gender}")

        # 验证班级
        class_name = str(row.get('班级', '')).strip()
        if not class_name:
            errors.append("班级不能为空")
        elif len(class_name) < 2 or len(class_name) > 20:
            errors.append(f"班级名称长度应在2-20个字符之间，当前长度: {len(class_name)}")

        # 验证入学年份
        enrollment_year = row.get('入学年份')
        if enrollment_year is None or str(enrollment_year).strip() == '':
            errors.append("入学年份不能为空")
        else:
            try:
                enrollment_year = int(enrollment_year)
                if enrollment_year < 2000 or enrollment_year > 2100:
                    errors.append(f"入学年份应在2000-2100之间，当前值: {enrollment_year}")
            except (ValueError, TypeError):
                errors.append(f"入学年份格式错误，应为整数，当前值: {enrollment_year}")

        if errors:
            return {
                'data': None,
                'errors': errors
            }

        return {
            'data': {
                'student_id': student_id,
                'name': name,
                'gender': gender,
                'class_name': class_name,
                'enrollment_year': int(enrollment_year)
            },
            'errors': []
        }

    def batch_import_students(
        self, 
        students_data: list[dict], 
        operator_id: Optional[int] = None,
        operator_name: Optional[str] = None
    ) -> dict:
        """
        批量导入学生数据
        
        Args:
            students_data: 学生数据列表
            operator_id: 操作用户ID
            operator_name: 操作用户名
            
        Returns:
            导入结果统计
        """
        total_rows = len(students_data)
        success_count = 0
        fail_count = 0
        errors = []

        # 检查重复学号（在文件内）
        seen_student_ids = set()
        duplicate_in_file = []
        
        for idx, student in enumerate(students_data, start=2):
            student_id = student['student_id']
            if student_id in seen_student_ids:
                duplicate_in_file.append({
                    'row': idx,
                    'student_id': student_id,
                    'error': '文件中存在重复学号'
                })
            seen_student_ids.add(student_id)

        if duplicate_in_file:
            raise ValidationException(
                f"文件中存在重复学号: {[d['student_id'] for d in duplicate_in_file]}"
            )

        # 逐条导入
        for idx, student_data in enumerate(students_data, start=2):
            try:
                # 检查学号是否已存在于数据库
                existing = self.student_repo.get_by_student_id(student_data['student_id'])
                if existing:
                    errors.append({
                        'row': idx,
                        'student_id': student_data['student_id'],
                        'field': '学号',
                        'error': '学号已存在',
                        'value': student_data['student_id']
                    })
                    fail_count += 1
                    continue

                # 创建学生记录
                self.student_repo.create(student_data)
                success_count += 1

            except Exception as e:
                errors.append({
                    'row': idx,
                    'student_id': student_data.get('student_id', '未知'),
                    'field': '数据库',
                    'error': str(e),
                    'value': ''
                })
                fail_count += 1

        # 提交事务
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            raise ValidationException(f"数据库提交失败: {str(e)}")

        return {
            'total_rows': total_rows,
            'success_count': success_count,
            'fail_count': fail_count,
            'errors': errors
        }

    def import_from_file(
        self,
        file_content: bytes,
        filename: str,
        operator_id: Optional[int] = None,
        operator_name: Optional[str] = None
    ) -> dict:
        """
        从文件导入学生数据
        
        Args:
            file_content: 文件内容
            filename: 文件名
            operator_id: 操作用户ID
            operator_name: 操作用户名
            
        Returns:
            导入结果统计
        """
        # 根据文件扩展名选择解析器
        if filename.endswith('.csv'):
            result = self.parse_csv_file(file_content)
        elif filename.endswith('.xlsx'):
            result = self.parse_excel_file(file_content)
        else:
            raise ValidationException("不支持的文件格式，请上传 .xlsx 或 .csv 文件")

        students_data = result['students']
        parse_errors = result['errors']

        if not students_data and not parse_errors:
            raise ValidationException("文件中没有有效的学生数据")

        if len(students_data) > 1000:
            raise ValidationException("单次导入不能超过1000条记录")

        # 合并解析错误和导入结果
        import_result = self.batch_import_students(
            students_data, 
            operator_id=operator_id,
            operator_name=operator_name
        )
        
        # 将解析错误添加到导入结果中
        import_result['errors'] = parse_errors + import_result.get('errors', [])
        import_result['fail_count'] = import_result['fail_count'] + len(parse_errors)
        
        return import_result

    def preview_file(self, file_content: bytes, filename: str) -> dict:
        """
        预览文件中的学生数据（不实际导入）
        
        Args:
            file_content: 文件内容
            filename: 文件名
            
        Returns:
            预览结果
        """
        # 根据文件扩展名选择解析器
        if filename.endswith('.csv'):
            result = self.parse_csv_file(file_content)
        elif filename.endswith('.xlsx'):
            result = self.parse_excel_file(file_content)
        else:
            raise ValidationException("不支持的文件格式，请上传 .xlsx 或 .csv 文件")

        students_data = result['students']
        parse_errors = result['errors']

        if not students_data and not parse_errors:
            raise ValidationException("文件中没有有效的学生数据")

        # 验证每条记录并生成预览
        preview = []
        valid_count = 0
        invalid_count = 0
        errors = list(parse_errors)  # 复制解析错误

        # 添加解析错误的行到预览
        for error in parse_errors:
            preview.append({
                'row': error['row'],
                'student_id': error.get('student_id', '未知'),
                'name': '未知',
                'gender': '未知',
                'class_name': '未知',
                'enrollment_year': 0,
                'status': 'invalid',
                'errors': [error['error']]
            })
            invalid_count += 1

        # 检查有效的学生数据
        for idx, student in enumerate(students_data, start=2):
            try:
                # 检查学号是否已存在
                existing = self.student_repo.get_by_student_id(student['student_id'])
                if existing:
                    preview.append({
                        'row': idx,
                        'student_id': student['student_id'],
                        'name': student['name'],
                        'gender': student['gender'],
                        'class_name': student['class_name'],
                        'enrollment_year': student['enrollment_year'],
                        'status': 'invalid',
                        'errors': ['学号已存在']
                    })
                    invalid_count += 1
                    errors.append({
                        'row': idx,
                        'student_id': student['student_id'],
                        'field': '学号',
                        'error': '学号已存在',
                        'value': student['student_id']
                    })
                else:
                    preview.append({
                        'row': idx,
                        'student_id': student['student_id'],
                        'name': student['name'],
                        'gender': student['gender'],
                        'class_name': student['class_name'],
                        'enrollment_year': student['enrollment_year'],
                        'status': 'valid'
                    })
                    valid_count += 1
            except Exception as e:
                preview.append({
                    'row': idx,
                    'student_id': student.get('student_id', '未知'),
                    'name': student.get('name', '未知'),
                    'gender': student.get('gender', '未知'),
                    'class_name': student.get('class_name', '未知'),
                    'enrollment_year': student.get('enrollment_year', 0),
                    'status': 'invalid',
                    'errors': [str(e)]
                })
                invalid_count += 1
                errors.append({
                    'row': idx,
                    'student_id': student.get('student_id', '未知'),
                    'field': '数据',
                    'error': str(e),
                    'value': ''
                })

        return {
            'total_rows': len(students_data) + len(parse_errors),
            'valid_rows': valid_count,
            'invalid_rows': invalid_count,
            'preview': preview,
            'errors': errors
        }

    def generate_template(self, format: str = 'xlsx') -> bytes:
        """
        生成导入模板
        
        Args:
            format: 模板格式 ('xlsx' 或 'csv')
            
        Returns:
            模板文件内容
        """
        if format == 'csv':
            return self._generate_csv_template()
        elif format == 'xlsx':
            return self._generate_excel_template()
        else:
            raise ValidationException("不支持的模板格式，请选择 'xlsx' 或 'csv'")

    def _generate_csv_template(self) -> bytes:
        """生成 CSV 模板"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['学号', '姓名', '性别', '班级', '入学年份'])
        
        # 写入示例数据
        writer.writerow(['20260001', '张三', '男', '三年一班', '2026'])
        writer.writerow(['20260002', '李四', '女', '三年一班', '2026'])
        
        return output.getvalue().encode('utf-8-sig')  # 使用 utf-8-sig 以支持 Excel 打开

    def _generate_excel_template(self) -> bytes:
        """生成 Excel 模板"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ValidationException("请安装 openpyxl 库以支持 Excel 模板生成")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生导入模板"

        # 定义样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ['学号', '姓名', '性别', '班级', '入学年份']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入示例数据
        example_data = [
            ['20260001', '张三', '男', '三年一班', 2026],
            ['20260002', '李四', '女', '三年一班', 2026],
        ]
        
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入说明行
        ws.cell(row=5, column=1, value="说明：")
        ws.cell(row=6, column=1, value="1. 学号：8位数字，格式为年份+4位序号（如：20260001）")
        ws.cell(row=7, column=1, value="2. 姓名：2-20个字符")
        ws.cell(row=8, column=1, value="3. 性别：男 或 女")
        ws.cell(row=9, column=1, value="4. 班级：2-20个字符（如：三年一班）")
        ws.cell(row=10, column=1, value="5. 入学年份：2000-2100之间的整数")

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    # ==================== 成绩导入相关方法 ====================

    def parse_grade_csv_file(self, file_content: bytes, exam_type: str, exam_date: str) -> dict:
        """
        解析成绩 CSV 文件内容
        
        Args:
            file_content: CSV 文件内容（字节）
            exam_type: 考试类型
            exam_date: 考试日期
            
        Returns:
            包含 grades 和 errors 的字典
        """
        try:
            # 尝试不同编码，优先使用 utf-8-sig 处理 BOM
            content = None
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
                try:
                    content = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                raise ValidationException("无法识别文件编码，请使用 UTF-8 或 GBK 编码")

            # 去除可能残留的 BOM 字符
            if content.startswith('\ufeff'):
                content = content[1:]

            # 解析 CSV
            reader = csv.DictReader(io.StringIO(content))
            grades = []
            errors = []
            
            for row_num, row in enumerate(reader, start=2):
                # 清理字段名和值的空白字符
                cleaned_row = {k.strip(): v for k, v in row.items() if k is not None}
                result = self._validate_grade_row(cleaned_row, row_num, exam_type, exam_date)
                if result['data']:
                    grades.append(result['data'])
                if result['errors']:
                    errors.append({
                        'row': row_num,
                        'student_id': str(cleaned_row.get('学号', '')).strip() or '未知',
                        'error': '; '.join(result['errors'])
                    })
            
            return {'grades': grades, 'errors': errors}

        except csv.Error as e:
            raise ValidationException(f"CSV 文件格式错误: {str(e)}")

    def parse_grade_excel_file(self, file_content: bytes, exam_type: str, exam_date: str) -> dict:
        """
        解析成绩 Excel 文件内容
        
        Args:
            file_content: Excel 文件内容（字节）
            exam_type: 考试类型
            exam_date: 考试日期
            
        Returns:
            包含 grades 和 errors 的字典
        """
        try:
            import openpyxl
        except ImportError:
            raise ValidationException("请安装 openpyxl 库以支持 Excel 文件")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active

            if ws is None:
                raise ValidationException("Excel 文件没有活动工作表")

            # 读取表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")

            # 验证表头
            required_headers = ['学号', '科目', '分数']
            for header in required_headers:
                if header not in headers:
                    raise ValidationException(f"Excel 文件缺少必要列: {header}")

            # 解析数据行
            grades = []
            errors = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                    
                row_dict = dict(zip(headers, row))
                result = self._validate_grade_row(row_dict, row_num, exam_type, exam_date)
                if result['data']:
                    grades.append(result['data'])
                if result['errors']:
                    errors.append({
                        'row': row_num,
                        'student_id': str(row_dict.get('学号', '')).strip() or '未知',
                        'error': '; '.join(result['errors'])
                    })

            wb.close()
            return {'grades': grades, 'errors': errors}

        except Exception as e:
            if isinstance(e, ValidationException):
                raise
            raise ValidationException(f"Excel 文件解析错误: {str(e)}")

    def _validate_grade_row(self, row: dict, row_num: int, exam_type: str, exam_date: str) -> dict:
        """
        验证单行成绩数据
        
        Args:
            row: 行数据字典
            row_num: 行号
            exam_type: 考试类型
            exam_date: 考试日期
            
        Returns:
            包含 data 和 errors 的字典
        """
        errors = []

        # 验证学号
        student_id = str(row.get('学号', '')).strip()
        if not student_id:
            errors.append("学号不能为空")
        elif len(student_id) != 8 or not student_id.isdigit():
            errors.append(f"学号格式错误，应为8位数字，当前值: {student_id}")

        # 验证科目
        from src.core.constants import SUBJECTS
        subject = str(row.get('科目', '')).strip()
        if not subject:
            errors.append("科目不能为空")
        elif subject not in SUBJECTS:
            errors.append(f"科目'{subject}'不在允许范围内，有效科目: {', '.join(SUBJECTS)}")

        # 验证分数
        score = row.get('分数')
        if score is None or str(score).strip() == '':
            errors.append("分数不能为空")
        else:
            try:
                score = float(score)
                if score < 0 or score > 100:
                    errors.append(f"分数应在0-100之间，当前值: {score}")
            except (ValueError, TypeError):
                errors.append(f"分数格式错误，应为数字，当前值: {score}")

        # 验证总分（可选）
        total_score = None
        total_score_val = row.get('总分')
        if total_score_val is not None and str(total_score_val).strip() != '':
            try:
                total_score = float(total_score_val)
                if total_score < 0:
                    errors.append(f"总分不能为负数，当前值: {total_score}")
            except (ValueError, TypeError):
                errors.append(f"总分格式错误，应为数字，当前值: {total_score_val}")

        # 验证考试类型（如果文件中有则使用文件中的，否则使用参数）
        row_exam_type = str(row.get('考试类型', '')).strip()
        if row_exam_type:
            exam_type = row_exam_type

        # 验证考试日期（如果文件中有则使用文件中的，否则使用参数）
        row_exam_date = str(row.get('考试日期', '')).strip()
        if row_exam_date:
            exam_date = row_exam_date

        # 验证日期格式
        try:
            from datetime import datetime, date
            exam_date_obj = datetime.strptime(exam_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            errors.append(f"考试日期格式错误，应为 YYYY-MM-DD，当前值: {exam_date}")
            exam_date_obj = None

        if errors:
            return {
                'data': None,
                'errors': errors
            }

        data = {
            'student_id': student_id,
            'subject': subject,
            'score': float(score),
            'exam_type': exam_type,
            'exam_date': exam_date_obj,
        }
        if total_score is not None:
            data['total_score'] = total_score

        return {
            'data': data,
            'errors': []
        }

    def import_grades_from_file(
        self,
        file_content: bytes,
        filename: str,
        exam_type: str,
        exam_date: str,
        operator_id: Optional[int] = None,
        operator_name: Optional[str] = None
    ) -> dict:
        """
        从文件导入成绩数据
        
        Args:
            file_content: 文件内容
            filename: 文件名
            exam_type: 考试类型
            exam_date: 考试日期
            operator_id: 操作用户ID
            operator_name: 操作用户名
            
        Returns:
            导入结果统计
        """
        # 根据文件扩展名选择解析器
        if filename.endswith('.csv'):
            result = self.parse_grade_csv_file(file_content, exam_type, exam_date)
        elif filename.endswith('.xlsx'):
            result = self.parse_grade_excel_file(file_content, exam_type, exam_date)
        else:
            raise ValidationException("不支持的文件格式，请上传 .xlsx 或 .csv 文件")

        grades_data = result['grades']
        parse_errors = result['errors']

        if not grades_data and not parse_errors:
            raise ValidationException("文件中没有有效的成绩数据")

        if len(grades_data) > 5000:
            raise ValidationException("单次导入不能超过5000条记录")

        return self.batch_import_grades(
            grades_data,
            operator_id=operator_id,
            operator_name=operator_name
        )

    def batch_import_grades(
        self,
        grades_data: list[dict],
        operator_id: Optional[int] = None,
        operator_name: Optional[str] = None
    ) -> dict:
        """
        批量导入成绩数据
        
        Args:
            grades_data: 成绩数据列表
            operator_id: 操作用户ID
            operator_name: 操作用户名
            
        Returns:
            导入结果统计
        """
        from src.models.grade import Grade

        total_rows = len(grades_data)
        success_count = 0
        fail_count = 0
        errors = []
        success_items = []
        failed_items = []

        for idx, grade_data in enumerate(grades_data, start=2):
            try:
                # 检查学生是否存在
                student = self.student_repo.get_by_student_id(grade_data['student_id'])
                if not student:
                    errors.append({
                        'row': idx,
                        'student_id': grade_data['student_id'],
                        'field': '学号',
                        'error': '学号不存在',
                        'value': grade_data['student_id']
                    })
                    failed_items.append({
                        'row': idx,
                        'student_id': grade_data['student_id'],
                        'subject': grade_data['subject'],
                        'score': grade_data['score'],
                        'error': '学号不存在'
                    })
                    fail_count += 1
                    continue

                # 检查成绩是否已存在（唯一约束：学号+科目+考试类型）
                existing = self.db.query(Grade).filter(
                    Grade.student_id == grade_data['student_id'],
                    Grade.subject == grade_data['subject'],
                    Grade.exam_type == grade_data['exam_type']
                ).first()

                if existing:
                    # 更新已有成绩
                    existing.score = grade_data['score']
                    existing.exam_date = grade_data['exam_date']
                    success_count += 1
                    success_items.append({
                        'row': idx,
                        'student_id': grade_data['student_id'],
                        'name': student.name,
                        'subject': grade_data['subject'],
                        'score': grade_data['score'],
                        'action': 'updated'
                    })
                else:
                    # 创建新成绩
                    new_grade = Grade(
                        student_id=grade_data['student_id'],
                        subject=grade_data['subject'],
                        score=grade_data['score'],
                        exam_type=grade_data['exam_type'],
                        exam_date=grade_data['exam_date']
                    )
                    self.db.add(new_grade)
                    success_count += 1
                    success_items.append({
                        'row': idx,
                        'student_id': grade_data['student_id'],
                        'name': student.name,
                        'subject': grade_data['subject'],
                        'score': grade_data['score'],
                        'action': 'created'
                    })

            except Exception as e:
                errors.append({
                    'row': idx,
                    'student_id': grade_data.get('student_id', '未知'),
                    'field': '数据库',
                    'error': str(e),
                    'value': ''
                })
                failed_items.append({
                    'row': idx,
                    'student_id': grade_data.get('student_id', '未知'),
                    'subject': grade_data.get('subject', '未知'),
                    'score': grade_data.get('score', 0),
                    'error': str(e)
                })
                fail_count += 1

        # 提交事务
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            raise ValidationException(f"数据库提交失败: {str(e)}")

        # 保存总分（从成绩数据中提取每个学生的总分）
        self._save_exam_totals(grades_data)

        return {
            'total_rows': total_rows,
            'success_count': success_count,
            'fail_count': fail_count,
            'errors': errors,
            'success_items': success_items,
            'failed_items': failed_items
        }

    def _save_exam_totals(self, grades_data: list[dict]) -> None:
        """
        保存学生考试总分

        Args:
            grades_data: 成绩数据列表
        """
        # 按学生+考试类型+考试日期分组，提取总分
        total_map: dict[str, dict] = {}
        for grade_data in grades_data:
            student_id = grade_data['student_id']
            exam_type = grade_data['exam_type']
            exam_date = grade_data['exam_date']
            total_score = grade_data.get('total_score')

            if total_score is not None:
                key = f"{student_id}_{exam_type}_{exam_date}"
                if key not in total_map:
                    total_map[key] = {
                        'student_id': student_id,
                        'exam_type': exam_type,
                        'exam_date': exam_date,
                        'total_score': total_score,
                    }

        # 保存总分记录
        for key, data in total_map.items():
            try:
                existing = self.db.query(StudentExamTotal).filter(
                    StudentExamTotal.student_id == data['student_id'],
                    StudentExamTotal.exam_type == data['exam_type'],
                    StudentExamTotal.exam_date == data['exam_date'],
                ).first()

                if existing:
                    existing.total_score = data['total_score']
                else:
                    total_record = StudentExamTotal(
                        student_id=data['student_id'],
                        exam_type=data['exam_type'],
                        exam_date=data['exam_date'],
                        total_score=data['total_score'],
                    )
                    self.db.add(total_record)
            except Exception:
                pass  # 总分保存失败不影响主流程

        try:
            self.db.commit()
        except Exception:
            self.db.rollback()

    def generate_grade_template(self, format: str = 'xlsx') -> bytes:
        """
        生成成绩导入模板
        
        Args:
            format: 模板格式 ('xlsx' 或 'csv')
            
        Returns:
            模板文件内容
        """
        if format == 'csv':
            return self._generate_grade_csv_template()
        elif format == 'xlsx':
            return self._generate_grade_excel_template()
        else:
            raise ValidationException("不支持的模板格式，请选择 'xlsx' 或 'csv'")

    def _generate_grade_csv_template(self) -> bytes:
        """生成成绩 CSV 模板"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['学号', '科目', '考试类型', '分数', '考试日期'])
        
        # 写入示例数据
        writer.writerow(['20260001', '数学', '期中', '95.0', '2026-04-15'])
        writer.writerow(['20260001', '语文', '期中', '88.5', '2026-04-15'])
        writer.writerow(['20260001', '英语', '期中', '92.0', '2026-04-15'])
        
        return output.getvalue().encode('utf-8-sig')

    def _generate_grade_excel_template(self) -> bytes:
        """生成成绩 Excel 模板"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ValidationException("请安装 openpyxl 库以支持 Excel 模板生成")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"

        # 定义样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ['学号', '科目', '考试类型', '分数', '考试日期']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入示例数据
        example_data = [
            ['20260001', '数学', '期中', 95.0, '2026-04-15'],
            ['20260001', '语文', '期中', 88.5, '2026-04-15'],
            ['20260001', '英语', '期中', 92.0, '2026-04-15'],
        ]
        
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入说明行
        ws.cell(row=6, column=1, value="说明：")
        ws.cell(row=7, column=1, value="1. 学号：8位数字，必须是已存在的学生")
        ws.cell(row=8, column=1, value="2. 科目：如数学、语文、英语等")
        ws.cell(row=9, column=1, value="3. 考试类型：期中、期末、月考、单元测试")
        ws.cell(row=10, column=1, value="4. 分数：0-100之间的数字，支持小数")
        ws.cell(row=11, column=1, value="5. 考试日期：格式为 YYYY-MM-DD（如：2026-04-15）")

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
